"""
Веб-интерфейс для парсера Яндекс рекламы
"""

from flask import Flask, render_template, request, jsonify, send_file
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import json
import csv
from datetime import datetime
import os
import threading
import logging
import re
from urllib.parse import quote
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
import xml.etree.ElementTree as ET
from xml.dom import minidom

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Глобальные переменные для отслеживания статуса
parsing_status = {
    'is_running': False,
    'progress': 0,
    'message': '',
    'total_ads': 0,
    'current_query': ''
}

results = []

# Настройки приложения
SETTINGS_FILE = 'settings.json'
app_settings = {
    'premium_domains': [],
    'max_pages': 2,
    'headless': False,
    'last_queries': []
}

# Загрузка настроек при старте
def load_settings():
    """Загрузка настроек из файла"""
    global app_settings
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                app_settings = json.load(f)
                logger.info(f"Настройки загружены: {len(app_settings.get('premium_domains', []))} премиум-доменов")
    except Exception as e:
        logger.error(f"Ошибка загрузки настроек: {e}")

def save_settings():
    """Сохранение настроек в файл"""
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(app_settings, f, ensure_ascii=False, indent=2)
        logger.info("Настройки сохранены")
        return True
    except Exception as e:
        logger.error(f"Ошибка сохранения настроек: {e}")
        return False


class YandexAdParser:
    def __init__(self, headless=True, status_callback=None, debug_descriptions=False):
        self.chrome_options = Options()
        if headless:
            self.chrome_options.add_argument("--headless")
        
        self.chrome_options.add_argument("--no-sandbox")
        self.chrome_options.add_argument("--disable-dev-shm-usage")
        self.chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        self.chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.chrome_options.add_experimental_option('useAutomationExtension', False)
        self.chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        self.debug_descriptions = debug_descriptions  # Режим отладки описаний
        
        # Добавляем preferences для обхода обнаружения и блокировки попапов
        self.chrome_options.add_experimental_option("prefs", {
            "profile.default_content_setting_values.notifications": 2,  # Блокировка уведомлений
            "profile.default_content_setting_values.popups": 0,         # Блокировка попапов
            "profile.default_content_setting_values.geolocation": 2,    # Блокировка геолокации
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.default_content_setting_values.media_stream": 2,   # Блокировка камеры/микрофона
        })
        
        self.driver = None
        self.results = []
        self.status_callback = status_callback
        self.headless = headless
    
    def update_status(self, message):
        """Обновление статуса парсинга"""
        if self.status_callback:
            self.status_callback(message)
        logger.info(message)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
    
    def close_popups(self):
        """Универсальная функция для закрытия всплывающих окон"""
        try:
            # Список селекторов для кнопок закрытия попапов
            close_selectors = [
                # Яндекс специфичные
                "button[aria-label='Закрыть']",
                "button[aria-label='Close']",
                ".modal-close",
                ".popup-close",
                ".close-button",
                "button.close",
                "button[class*='close']",
                "button[class*='Close']",
                # Общие паттерны
                "div[class*='modal'] button[class*='close']",
                "div[class*='popup'] button[class*='close']",
                "div[class*='dialog'] button[class*='close']",
                "[class*='Modal'] [class*='Close']",
                "[class*='Popup'] [class*='Close']",
                # SVG иконки закрытия
                "button svg[class*='close']",
                "button svg[class*='cross']",
                # Яндекс баннеры и уведомления
                ".banner__close",
                ".notification__close",
                ".promo__close",
                # Кнопки с текстом
                "button:contains('Закрыть')",
                "button:contains('Понятно')",
                "button:contains('OK')",
                "a[class*='close']"
            ]
            
            closed_count = 0
            for selector in close_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            try:
                                element.click()
                                closed_count += 1
                                logger.debug(f"Закрыт попап (селектор: {selector})")
                                time.sleep(0.3)  # Небольшая пауза после закрытия
                            except:
                                # Пробуем через JavaScript
                                try:
                                    self.driver.execute_script("arguments[0].click();", element)
                                    closed_count += 1
                                    logger.debug(f"Закрыт попап через JS (селектор: {selector})")
                                    time.sleep(0.3)
                                except:
                                    pass
                except:
                    continue
            
            # Закрываем оверлеи (затемнения)
            try:
                overlays = self.driver.find_elements(By.CSS_SELECTOR, 
                    "div[class*='overlay'], div[class*='backdrop'], div[class*='mask']")
                for overlay in overlays:
                    if overlay.is_displayed():
                        try:
                            self.driver.execute_script("arguments[0].remove();", overlay)
                            closed_count += 1
                            logger.debug("Удален оверлей")
                        except:
                            pass
            except:
                pass
            
            if closed_count > 0:
                logger.info(f"Закрыто всплывающих окон: {closed_count}")
            
            return closed_count
            
        except Exception as e:
            logger.debug(f"Ошибка при закрытии попапов: {e}")
            return 0
    
    def start_browser(self):
        """Запуск браузера"""
        try:
            logger.info("Инициализация Chrome WebDriver...")
            self.driver = webdriver.Chrome(options=self.chrome_options)
            self.driver.maximize_window()
            logger.info("Браузер успешно запущен")
            
            # Скрываем признаки автоматизации
            logger.info("Применение anti-detection мер...")
            self.driver.execute_cdp_cmd('Network.setUserAgentOverride', {
                "userAgent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            })
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            self.update_status("Браузер запущен")
            
            # Открываем Яндекс для генерации токенов и обхода капчи
            if not self.headless:
                logger.info("Открытие Яндекс для инициализации сессии...")
                self.update_status("Открываю Яндекс для инициализации сессии...")
                self.driver.get("https://yandex.ru")
                time.sleep(3)
                
                # Автоматическое закрытие модалки Яндекса о расширении
                try:
                    logger.info("Проверка и закрытие модального окна...")
                    close_selectors = [
                        "button[aria-label='Закрыть']",
                        "button.close",
                        ".modal-close",
                        "button[class*='close']",
                        "div[class*='modal'] button",
                        "[class*='popup'] button[class*='close']"
                    ]
                    
                    for selector in close_selectors:
                        try:
                            close_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                            if close_button.is_displayed():
                                close_button.click()
                                logger.info(f"Модальное окно закрыто (селектор: {selector})")
                                time.sleep(1)
                                break
                        except:
                            continue
                except Exception as e:
                    logger.debug(f"Модальное окно не найдено или уже закрыто: {e}")
                
                logger.info("Сессия инициализирована")
                self.update_status("Сессия инициализирована. Если появилась капча - решите её вручную.")
                time.sleep(5)  # Даём время решить капчу если она есть
            
            return True
        except Exception as e:
            error_msg = f"Ошибка при запуске браузера: {e}"
            logger.error(error_msg)
            self.update_status(error_msg)
            return False
    
    def parse_yandex_ads(self, query, max_pages=3):
        """Парсинг рекламы из Яндекса"""
        logger.info(f"Начало парсинга по запросу: '{query}'")
        self.update_status(f"Начинаю парсинг по запросу: '{query}'")
        
        try:
            # Формируем URL напрямую (как в примере Яндекса)
            encoded_query = quote(query)
            search_url = f"https://yandex.ru/search/?text={encoded_query}&lr=213"
            
            logger.info(f"Сформирован URL: {search_url}")
            self.update_status(f"Переход на страницу поиска...")
            
            self.driver.get(search_url)
            time.sleep(3)  # Пауза для загрузки
            
            # Автоматическое закрытие всплывающих окон
            logger.info("Закрытие всплывающих окон...")
            self.close_popups()
            time.sleep(2)
            
            # Закрытие модалок на странице поиска (устаревший код - оставляем для совместимости)
            try:
                close_selectors = [
                    "button[aria-label='Закрыть']",
                    "button.close",
                    ".modal-close",
                    "button[class*='close']"
                ]
                for selector in close_selectors:
                    try:
                        close_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                        if close_button.is_displayed():
                            close_button.click()
                            time.sleep(1)
                            break
                    except:
                        continue
            except:
                pass
            
            # Проверка на капчу
            current_url = self.driver.current_url
            logger.info(f"Текущий URL: {current_url}")
            
            if "showcaptcha" in current_url or "капча" in self.driver.page_source.lower():
                warning_msg = "Обнаружена капча! Решите её вручную в открытом браузере."
                logger.warning(warning_msg)
                self.update_status(warning_msg)
                if not self.headless:
                    self.update_status("Ожидание решения капчи... (30 сек)")
                    time.sleep(30)
                    # Проверяем, решена ли капча
                    if "showcaptcha" in self.driver.current_url:
                        logger.error("Капча не решена")
                        self.update_status("Капча не решена. Пропускаю запрос.")
                        return
                    else:
                        logger.info("Капча решена, продолжаем")
                        self.update_status("Капча решена, продолжаем парсинг")
                else:
                    error_msg = "Невозможно решить капчу в headless режиме. Используйте режим с браузером."
                    logger.error(error_msg)
                    self.update_status(error_msg)
                    return
            
            for page in range(max_pages):
                logger.info(f"Обработка страницы {page + 1} из {max_pages}")
                self.update_status(f"Обработка страницы {page + 1} из {max_pages}")
                
                # Закрываем попапы перед обработкой страницы
                self.close_popups()
                
                # Сначала пытаемся найти рекламу
                ads_found_before = len(self.results)
                
                self._parse_yandex_all_ads(query)
                
                ads_found_after = len(self.results)
                new_ads = ads_found_after - ads_found_before
                
                logger.info(f"На странице {page + 1} найдено новых объявлений: {new_ads}")
                
                if page < max_pages - 1:
                    try:
                        logger.info("Поиск кнопки следующей страницы...")
                        # Ищем разные варианты кнопки "Следующая"
                        next_selectors = [
                            "a.Pager-Item_type_next",
                            "a[aria-label='Следующая страница']",
                            "div.pager__item_kind_next a",
                            "a.link_theme_normal[href*='p=']"
                        ]
                        
                        next_button = None
                        for selector in next_selectors:
                            try:
                                next_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                                if next_button and next_button.is_displayed():
                                    logger.info(f"Кнопка найдена по селектору: {selector}")
                                    break
                            except:
                                continue
                        
                        if next_button:
                            logger.info("Переход на следующую страницу")
                            self.driver.execute_script("arguments[0].scrollIntoView();", next_button)
                            time.sleep(1)
                            next_button.click()
                            time.sleep(3)
                            
                            # Закрываем попапы после перехода на новую страницу
                            logger.info("Закрытие всплывающих окон на новой странице...")
                            self.close_popups()
                            time.sleep(2)
                        else:
                            logger.warning("Кнопка следующей страницы не найдена")
                            self.update_status("Больше нет страниц")
                            break
                    except Exception as e:
                        logger.warning(f"Ошибка при переходе на следующую страницу: {e}")
                        self.update_status("Больше нет страниц")
                        break
            
            found_count = len([r for r in self.results if r['query'] == query])
            logger.info(f"Парсинг завершен. Найдено объявлений: {found_count}")
            self.update_status(f"Найдено {found_count} объявлений")
            
        except Exception as e:
            error_msg = f"Ошибка при парсинге: {e}"
            logger.error(error_msg, exc_info=True)
            self.update_status(error_msg)
    
    def _parse_yandex_all_ads(self, query):
        """Универсальный парсинг всей рекламы на странице"""
        try:
            logger.info("Начало поиска рекламных блоков...")
            
            # Ждем загрузки страницы
            time.sleep(2)
            
            # Ищем все возможные рекламные блоки
            ad_selectors = [
                # Прямые рекламные блоки Яндекс.Директ
                "li.serp-item[data-cid]",
                "div.OrganicTitle",
                "div.Organic",
                ".serp-item",
                # Дополнительные селекторы
                "[class*='direct']",
                "[class*='adv']",
                "[class*='commercial']"
            ]
            
            all_items = []
            for selector in ad_selectors:
                try:
                    items = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if items:
                        logger.info(f"Селектор '{selector}': найдено {len(items)} элементов")
                        all_items.extend(items)
                except Exception as e:
                    logger.debug(f"Селектор '{selector}' не сработал: {e}")
            
            logger.info(f"Всего элементов для обработки: {len(all_items)}")
            
            # Закрываем попапы перед обработкой элементов
            self.close_popups()
            
            # Обрабатываем найденные элементы
            processed_texts = set()
            ads_count = 0
            
            # Счетчики позиций для разных типов
            organic_position = 0
            direct_position = 0
            
            for item_idx, item in enumerate(all_items):
                # Периодически закрываем попапы (каждые 10 элементов)
                if item_idx > 0 and item_idx % 10 == 0:
                    self.close_popups()
                
                try:
                    # Проверяем, что элемент видим
                    if not item.is_displayed():
                        continue
                    
                    item_text = item.text.strip()
                    
                    # Пропускаем дубликаты и слишком короткие
                    if not item_text or len(item_text) < 10 or item_text in processed_texts:
                        continue
                    
                    processed_texts.add(item_text)
                    
                    ad_data = {
                        'platform': 'Яндекс',
                        'type': 'Реклама',
                        'query': query,
                        'title': '',
                        'description': '',
                        'url': '',
                        'display_url': '',
                        'preview_url': '',  # Новое поле для превью ссылки
                        'creative_image': '',
                        'position': 0,  # Позиция в выдаче (будет установлена позже)
                        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    
                    # Ищем заголовок - УЛУЧШЕННЫЙ МЕТОД
                    title_selectors = [
                        "h2 a", "h3 a", "a.Link_theme_outer", "a.OrganicTitle-Link",
                        "a.Organic-Title", ".OrganicTitle a", "a.link", "a[href]"
                    ]
                    
                    # Функция проверки, является ли текст URL-подобным
                    def is_url_like(text):
                        """Проверяет, похож ли текст на URL или путь"""
                        if not text:
                            return False
                        # Содержит символы типичные для URL
                        url_indicators = ['›', '/', '.ru', '.com', 'http', 'www']
                        return any(indicator in text.lower() for indicator in url_indicators)
                    
                    # Пытаемся найти заголовок
                    for t_sel in title_selectors:
                        try:
                            title_elem = item.find_element(By.CSS_SELECTOR, t_sel)
                            title_text = title_elem.text.strip()
                            
                            # Проверяем что это реальный заголовок, а не URL
                            if title_text and len(title_text) > 3 and not is_url_like(title_text):
                                ad_data['title'] = title_text
                                ad_data['url'] = title_elem.get_attribute('href') or ''
                                break
                        except:
                            continue
                    
                    # Если заголовок не найден, извлекаем первую НЕ-URL строку из текста
                    if not ad_data['title'] and item_text:
                        lines = [l.strip() for l in item_text.split('\n') if l.strip()]
                        for line in lines:
                            # Пропускаем метки рекламы
                            if line in ['Реклама', 'реклама', 'Промо', 'промо']:
                                continue
                            # Пропускаем URL-подобные строки
                            if is_url_like(line):
                                # Сохраняем как preview_url если ещё не сохранено
                                if not ad_data['preview_url'] and len(line) < 150:
                                    ad_data['preview_url'] = line
                                continue
                            # Нашли нормальный заголовок
                            if len(line) > 10 and len(line) < 200:
                                ad_data['title'] = line
                                break
                        
                        # Если всё ещё не нашли, берём первую строку
                        if not ad_data['title'] and lines:
                            ad_data['title'] = lines[0][:100]
                    
                    # Ищем описание - РАСШИРЕННАЯ СИСТЕМА с множественными стратегиями
                    desc_selectors = [
                        # ВЫСОКИЙ ПРИОРИТЕТ: Яндекс.Директ (рекламные блоки)
                        "div.Organic-Text",                          # Основное описание Директа
                        "div.Organic-TextBg div.Organic-Text",       # Описание в контейнере
                        "span.OrganicTextContentSpan",               # Текстовый контент
                        ".DirectSnippet-Text",                       # Текст сниппета Директа
                        ".DirectSnippet .text",                      # Альт. вариант
                        "div.text-container",                        # Общий контейнер текста
                        
                        # СРЕДНИЙ ПРИОРИТЕТ: Органические результаты
                        "div.Organic-TextBg",                        # Контейнер текста
                        "div[class*='Organic'][class*='TextBg']",    # Вариация контейнера
                        "div.TextContainer",                         # Общий контейнер
                        "div.SnippetText",                           # Текст сниппета
                        "div.organic__text",                         # Старая структура
                        "div.VanillaReact.Organic-Text",             # React компонент
                        "div.VanillaReact.Text",                     # Альт. React
                        
                        # СПЕЦИАЛЬНЫЕ селекторы для вложенных структур
                        ".Organic .text-container",
                        ".serp-item .text-container", 
                        ".Organic div[class*='Text']:not([class*='Title'])",
                        "div[class*='organic'] div[class*='text']",
                        
                        # НИЗКИЙ ПРИОРИТЕТ: Универсальные селекторы
                        "div[class*='Organic'][class*='Text']",      
                        "div[class*='snippet'][class*='text']",
                        "span[class*='Organic'][class*='Text']",
                        "div.Typo_text_m",
                        "div.Typo.Organic-Text",
                        "div.Typo.Text",
                        "span.Typo",
                        
                        # FALLBACK: Широкие селекторы (используются в последнюю очередь)
                        "div[class*='Text']:not([class*='Title']):not([class*='Url'])",
                        "div[class*='text']:not([class*='title']):not([class*='url'])",
                        "span[class*='Text']",
                        "div.Typo",
                        "div[class*='Snippet']",
                        "p",  # Параграфы
                        ".text"  # Класс .text
                    ]
                    
                    # Функция очистки текста описания
                    def clean_description_text(text, title):
                        """Очищает текст описания от шумов"""
                        if not text:
                            return ""
                        
                        # Убираем метки рекламы (должны быть в начале)
                        noise_labels = ['Реклама', 'реклама', 'Промо', 'промо', 'Promoted', 'Ad', 'Sponsored']
                        for label in noise_labels:
                            # Убираем только если это начало строки
                            if text.startswith(label):
                                text = text[len(label):].strip()
                        
                        # Убираем заголовок если он попал в описание
                        if title and text.startswith(title):
                            text = text[len(title):].strip()
                        
                        # Убираем URL (начинающиеся с http, https, www)
                        lines = text.split('\n')
                        cleaned_lines = []
                        for line in lines:
                            line = line.strip()
                            if not line.startswith(('http://', 'https://', 'www.', '//')):
                                cleaned_lines.append(line)
                        text = ' '.join(cleaned_lines)
                        
                        # Убираем множественные пробелы
                        text = ' '.join(text.split())
                        
                        return text.strip()
                    
                    # Пробуем найти описание по селекторам
                    description_found = False
                    for d_sel in desc_selectors:
                        if description_found:
                            break
                            
                        try:
                            desc_elems = item.find_elements(By.CSS_SELECTOR, d_sel)
                            
                            for desc_elem in desc_elems:
                                try:
                                    # Проверяем видимость элемента
                                    if not desc_elem.is_displayed():
                                        continue
                                    
                                    desc_text = desc_elem.text.strip()
                                    
                                    # Пропускаем слишком короткие
                                    if len(desc_text) < 15:
                                        continue
                                    
                                    # Пропускаем если это точно заголовок
                                    if desc_text == ad_data['title']:
                                        continue
                                    
                                    # Пропускаем если это URL
                                    if desc_text.startswith(('http://', 'https://', 'www.', '›', '/')):
                                        continue
                                    
                                    # Очищаем текст
                                    desc_text = clean_description_text(desc_text, ad_data['title'])
                                    
                                    # Проверяем что после очистки осталось что-то полезное
                                    if desc_text and len(desc_text) > 20:
                                        ad_data['description'] = desc_text[:500]
                                        description_found = True
                                        if self.debug_descriptions:
                                            logger.info(f"✓ Описание найдено через селектор: {d_sel}")
                                        break
                                except Exception as e:
                                    if self.debug_descriptions:
                                        logger.debug(f"Ошибка обработки элемента {d_sel}: {e}")
                                    continue
                        except Exception as e:
                            if self.debug_descriptions:
                                logger.debug(f"Селектор {d_sel} не сработал: {e}")
                            continue
                    
                    # МЕТОД 2: Умное извлечение из полного текста элемента (улучшенное)
                    if not ad_data['description']:
                        try:
                            full_text = item.text.strip()
                            if full_text and len(full_text) > 40:  # Снизили минимум
                                lines = [l.strip() for l in full_text.split('\n') if l.strip()]
                                
                                # Определяем стоп-слова
                                stop_keywords = ['Реклама', 'реклама', 'Промо', 'промо', 'Sponsored', 'Ad']
                                url_patterns = ['http://', 'https://', 'www.', '›', '//']
                                
                                description_lines = []
                                title_found = False
                                skip_next = False
                                
                                for i, line in enumerate(lines):
                                    line_stripped = line.strip()
                                    
                                    # Пропускаем пустые
                                    if not line_stripped:
                                        continue
                                    
                                    # Пропускаем метки рекламы (только точное совпадение)
                                    if line_stripped in stop_keywords:
                                        continue
                                    
                                    # Пропускаем URL
                                    if any(pattern in line_stripped for pattern in url_patterns):
                                        continue
                                    
                                    # Определяем заголовок
                                    if not title_found:
                                        if line_stripped == ad_data['title'] or (ad_data['title'] and ad_data['title'] in line_stripped):
                                            title_found = True
                                            continue
                                        # Если это первая строка и она не очень длинная - скорее всего это заголовок или метаданные
                                        if i == 0 and len(line_stripped) < 80:
                                            continue
                                    
                                    # Пропускаем очень короткие строки после заголовка (метаданные, даты и т.п.)
                                    if len(line_stripped) < 20:
                                        continue
                                    
                                    # Пропускаем строки, состоящие только из цифр, символов или коротких слов
                                    words = line_stripped.split()
                                    if len(words) < 3:  # Снизили минимум до 3 слов
                                        continue
                                    
                                    # Проверяем что это не просто набор цифр/символов
                                    alpha_count = sum(c.isalpha() for c in line_stripped)
                                    if alpha_count < len(line_stripped) * 0.5:  # Минимум 50% букв
                                        continue
                                    
                                    # Добавляем строку в описание
                                    description_lines.append(line_stripped)
                                    
                                    # Если накопили достаточно текста - останавливаемся
                                    joined_desc = ' '.join(description_lines)
                                    if len(joined_desc) > 80:  # Снизили минимум
                                        break
                                
                                if description_lines:
                                    final_desc = ' '.join(description_lines)
                                    # Финальная очистка
                                    final_desc = clean_description_text(final_desc, ad_data['title'])
                                    if len(final_desc) > 20:  # Снизили минимум
                                        ad_data['description'] = final_desc[:500]
                                        if self.debug_descriptions:
                                            logger.info(f"✓ Описание извлечено из полного текста")
                        except Exception as e:
                            if self.debug_descriptions:
                                logger.debug(f"Ошибка извлечения из полного текста: {e}")
                    
                    # МЕТОД 3: XPath - поиск текстовых блоков с оптимальной длиной
                    if not ad_data['description']:
                        try:
                            # Стратегия 1: Ищем div с прямым текстовым содержимым (не вложенным)
                            xpath_queries = [
                                # Текст непосредственно в div (наиболее вероятное описание)
                                ".//div[string-length(normalize-space(text())) >= 30]",
                                # Span с текстом
                                ".//span[string-length(normalize-space(text())) >= 30]",
                                # Любой элемент с достаточным текстом
                                ".//*[string-length(normalize-space(text())) >= 30 and string-length(normalize-space(text())) <= 500]",
                            ]
                            
                            candidates = []
                            
                            for xpath_query in xpath_queries:
                                try:
                                    text_elements = item.find_elements(By.XPATH, xpath_query)
                                    
                                    for elem in text_elements[:15]:  # Проверяем первые 15 кандидатов
                                        try:
                                            if not elem.is_displayed():
                                                continue
                                            
                                            # Получаем только прямой текст элемента (не вложенных детей)
                                            elem_text = self.driver.execute_script("""
                                                var element = arguments[0];
                                                var text = '';
                                                for (var i = 0; i < element.childNodes.length; i++) {
                                                    var node = element.childNodes[i];
                                                    if (node.nodeType === Node.TEXT_NODE) {
                                                        text += node.textContent;
                                                    }
                                                }
                                                return text.trim();
                                            """, elem)
                                            
                                            # Если прямого текста нет, берем весь текст элемента
                                            if not elem_text or len(elem_text) < 30:
                                                elem_text = elem.text.strip()
                                            
                                            # Базовые фильтры
                                            if not elem_text or len(elem_text) < 30:
                                                continue
                                            
                                            # Пропускаем заголовок
                                            if elem_text == ad_data['title']:
                                                continue
                                            
                                            # Пропускаем URL
                                            if elem_text.startswith(('http://', 'https://', 'www.', '//', '›')):
                                                continue
                                            
                                            # Пропускаем метки рекламы
                                            if elem_text in ['Реклама', 'реклама', 'Промо', 'промо', 'Ad', 'Sponsored']:
                                                continue
                                            
                                            # Вычисляем качество кандидата (больше = лучше)
                                            quality_score = 0
                                            
                                            # Длина в оптимальном диапазоне
                                            text_len = len(elem_text)
                                            if 50 <= text_len <= 300:
                                                quality_score += 10
                                            elif 30 <= text_len < 50:
                                                quality_score += 5
                                            elif text_len > 300:
                                                quality_score += 3
                                            
                                            # Содержит достаточно слов
                                            word_count = len(elem_text.split())
                                            if word_count >= 8:
                                                quality_score += 5
                                            elif word_count >= 5:
                                                quality_score += 3
                                            
                                            # Содержит кириллицу (русский текст)
                                            if re.search('[а-яА-ЯёЁ]', elem_text):
                                                quality_score += 3
                                            
                                            # Не содержит много цифр (не номер телефона и т.п.)
                                            digit_ratio = sum(c.isdigit() for c in elem_text) / max(len(elem_text), 1)
                                            if digit_ratio < 0.2:
                                                quality_score += 3
                                            
                                            # Не состоит только из заглавных букв
                                            upper_ratio = sum(c.isupper() for c in elem_text if c.isalpha()) / max(sum(c.isalpha() for c in elem_text), 1)
                                            if upper_ratio < 0.5:
                                                quality_score += 2
                                            
                                            # Содержит знаки препинания (признак связного текста)
                                            if any(p in elem_text for p in ['.', ',', '!', '?', ':', ';']):
                                                quality_score += 2
                                            
                                            candidates.append((quality_score, elem_text, xpath_query))
                                        except Exception as e:
                                            if self.debug_descriptions:
                                                logger.debug(f"Ошибка обработки XPath элемента: {e}")
                                            continue
                                except Exception as e:
                                    if self.debug_descriptions:
                                        logger.debug(f"XPath запрос не сработал: {e}")
                                    continue
                            
                            # Выбираем лучшего кандидата
                            if candidates:
                                candidates.sort(reverse=True, key=lambda x: x[0])
                                best_score, best_text, best_xpath = candidates[0]
                                
                                if best_score > 3:  # Снизили минимальный порог
                                    final_desc = clean_description_text(best_text, ad_data['title'])
                                    if len(final_desc) > 20:
                                        ad_data['description'] = final_desc[:500]
                                        if self.debug_descriptions:
                                            logger.info(f"✓ Описание найдено через XPath (качество: {best_score}, xpath: {best_xpath[:50]})")
                        except Exception as e:
                            if self.debug_descriptions:
                                logger.debug(f"Ошибка XPath поиска: {e}")
                    
                    # МЕТОД 4: Поиск по innerHTML (последняя попытка)
                    if not ad_data['description']:
                        try:
                            inner_html = item.get_attribute('innerHTML')
                            if inner_html:
                                # Убираем HTML теги
                                text_content = re.sub(r'<[^>]+>', ' ', inner_html)
                                text_content = re.sub(r'\s+', ' ', text_content).strip()
                                
                                if len(text_content) > 100:
                                    # Разбиваем на предложения
                                    sentences = re.split(r'[.!?]\s+', text_content)
                                    
                                    description_parts = []
                                    for sentence in sentences:
                                        sentence = sentence.strip()
                                        
                                        # Пропускаем короткие предложения
                                        if len(sentence) < 30:
                                            continue
                                        
                                        # Пропускаем если это заголовок
                                        if ad_data['title'] and ad_data['title'] in sentence:
                                            continue
                                        
                                        # Пропускаем URL
                                        if re.search(r'https?://|www\.', sentence):
                                            continue
                                        
                                        description_parts.append(sentence)
                                        
                                        # Ограничиваем количество предложений
                                        if len(' '.join(description_parts)) > 200:
                                            break
                                    
                                    if description_parts:
                                        final_desc = '. '.join(description_parts)
                                        final_desc = clean_description_text(final_desc, ad_data['title'])
                                        if len(final_desc) > 30:
                                            ad_data['description'] = final_desc[:500]
                                            if self.debug_descriptions:
                                                logger.info(f"✓ Описание извлечено из innerHTML")
                        except Exception as e:
                            if self.debug_descriptions:
                                logger.debug(f"Ошибка парсинга innerHTML: {e}")
                    
                    # ФИНАЛЬНАЯ ПРОВЕРКА и логирование
                    if not ad_data['description']:
                        # ПОСЛЕДНЯЯ ПОПЫТКА: используем item_text с минимальной обработкой
                        if item_text and len(item_text) > 30:
                            # Разбиваем на строки
                            lines = [l.strip() for l in item_text.split('\n') if l.strip()]
                            
                            # Ищем первую строку, которая не является заголовком и достаточно длинная
                            for line in lines:
                                # Пропускаем заголовок
                                if line == ad_data['title']:
                                    continue
                                
                                # Пропускаем URL
                                if line.startswith(('http', 'www', '›', '//')):
                                    continue
                                
                                # Пропускаем метки
                                if line in ['Реклама', 'реклама', 'Промо', 'промо']:
                                    continue
                                
                                # Если строка достаточно длинная - используем её
                                if len(line) >= 20:
                                    fallback_desc = clean_description_text(line, ad_data['title'])
                                    if len(fallback_desc) >= 15:  # Минимальный порог
                                        ad_data['description'] = fallback_desc[:500]
                                        if self.debug_descriptions:
                                            logger.info(f"⚠ Использован fallback (простой)")
                                        break
                        
                        # Если все еще нет описания - логируем для отладки
                        if not ad_data['description']:
                            item_class_check = item.get_attribute('class') or ''
                            logger.warning(f"❌ Описание не найдено: {ad_data['title'][:50]}...")
                            if self.debug_descriptions:
                                logger.debug(f"  Классы: {item_class_check[:150]}")
                                logger.debug(f"  Текст элемента (первые 200 символов): {item_text[:200] if item_text else 'Нет текста'}")
                                try:
                                    html_preview = item.get_attribute('outerHTML')[:400]
                                    logger.debug(f"  HTML: {html_preview}")
                                except:
                                    pass
                    else:
                        # Успешно извлекли описание
                        if self.debug_descriptions:
                            desc_preview = ad_data['description'][:60]
                            logger.info(f"✓✓✓ Описание успешно: {desc_preview}...")
                    
                    # Ищем отображаемый URL и превью ссылку
                    url_selectors = [
                        ".Organic-Url", "cite", ".path", ".url",
                        "[class*='url']", "[class*='domain']", ".Organic-Path"
                    ]
                    
                    for u_sel in url_selectors:
                        try:
                            url_elem = item.find_element(By.CSS_SELECTOR, u_sel)
                            url_text = url_elem.text.strip()
                            if url_text:
                                ad_data['display_url'] = url_text
                                # Если это похоже на путь с › (превью ссылка), сохраняем как preview_url
                                if '›' in url_text or '/' in url_text:
                                    ad_data['preview_url'] = url_text
                                break
                        except:
                            continue
                    
                    # Пытаемся найти креатив (изображение) - улучшенный парсинг
                    try:
                        # Расширенный список селекторов
                        img_selectors = [
                            'img',                              # Все изображения
                            '.thumb img',
                            '.image img',
                            '[class*="image"] img',
                            '[class*="thumb"] img',
                            '[class*="Image"] img',
                            '[class*="Thumb"] img',
                            '[class*="preview"] img',
                            '[class*="photo"] img',
                            '[class*="picture"] img',
                            'picture img',
                            '.Organic-Thumb img',
                            '.serp-item__thumb img',
                            'div[class*="Thumb"] img',
                            'img[src*="avatars"]',              # Яндекс аватарки
                            'img[src*="yandex"]'                # Яндекс картинки
                        ]
                        
                        for img_sel in img_selectors:
                            try:
                                img_elems = item.find_elements(By.CSS_SELECTOR, img_sel)
                                if not img_elems:
                                    continue
                                    
                                for img_elem in img_elems:
                                    try:
                                        # Пробуем разные атрибуты изображения
                                        img_src = (img_elem.get_attribute('src') or 
                                                 img_elem.get_attribute('data-src') or 
                                                 img_elem.get_attribute('data-lazy-src') or
                                                 img_elem.get_attribute('data-original') or
                                                 img_elem.get_attribute('srcset'))
                                        
                                        if img_src:
                                            # Если srcset, берем первый URL
                                            if 'srcset' in str(img_src):
                                                img_src = img_src.split(',')[0].split(' ')[0]
                                            
                                            # Проверяем что это реальный URL изображения
                                            if ('http' in img_src or img_src.startswith('//')):
                                                # Добавляем протокол если нужно
                                                if img_src.startswith('//'):
                                                    img_src = 'https:' + img_src
                                                
                                                # Пропускаем только совсем маленькие иконки
                                                skip_keywords = ['icon', 'favicon', 'avatar', '1x1.']
                                                if any(skip in img_src.lower() for skip in skip_keywords):
                                                    continue
                                                
                                                # Проверяем размер если доступен
                                                try:
                                                    width = img_elem.get_attribute('width')
                                                    height = img_elem.get_attribute('height')
                                                    
                                                    # Если размеры заданы и слишком маленькие - пропускаем
                                                    if width and height:
                                                        if int(width) < 40 or int(height) < 40:
                                                            continue
                                                except:
                                                    pass
                                                
                                                ad_data['creative_image'] = img_src
                                                logger.debug(f"✓ Найдено изображение: {img_src[:100]}")
                                                break
                                    except Exception as e:
                                        logger.debug(f"Ошибка обработки img элемента: {e}")
                                        continue
                                
                                if ad_data['creative_image']:
                                    break
                            except Exception as e:
                                logger.debug(f"Ошибка при поиске изображения с селектором {img_sel}: {e}")
                                continue
                    except Exception as e:
                        logger.debug(f"Общая ошибка при парсинге изображений: {e}")
                        pass
                    
                    # Функция проверки на кириллицу
                    def has_cyrillic(text):
                        return bool(re.search('[а-яА-ЯёЁ]', text))
                    
                    # Фильтруем кириллические URL
                    if ad_data['display_url'] and has_cyrillic(ad_data['display_url']):
                        # Пытаемся извлечь реальный URL из href
                        if ad_data['url']:
                            try:
                                from urllib.parse import urlparse
                                parsed = urlparse(ad_data['url'])
                                clean_domain = parsed.netloc or parsed.path.split('/')[0]
                                if clean_domain and not has_cyrillic(clean_domain):
                                    ad_data['display_url'] = clean_domain
                                else:
                                    ad_data['display_url'] = ''  # Не показываем кириллические URL
                            except:
                                ad_data['display_url'] = ''
                        else:
                            ad_data['display_url'] = ''
                    
                    # Определяем тип рекламы по классам или содержимому
                    item_class = item.get_attribute('class') or ''
                    item_html = item.get_attribute('outerHTML')[:500] or ''
                    
                    # Улучшенная классификация рекламы
                    is_premium = False
                    ad_type = 'Органика'
                    
                    # Проверяем соцсети
                    social_networks = ['vk.com', 'vk.ru', 'dzen.ru', 'ya.ru', 't.me', 'telegram', 'ok.ru', 'odnoklassniki.ru']
                    is_social = False
                    url_to_check = (ad_data['url'] or ad_data['display_url'] or '').lower()
                    
                    for social in social_networks:
                        if social in url_to_check:
                            is_social = True
                            ad_type = 'Соцсети'
                            break
                    
                    # Если не соцсеть, проверяем признаки рекламы
                    if not is_social:
                        # САМЫЙ НАДЕЖНЫЙ ПРИЗНАК: URL содержит yabs.yandex.ru (редирект Директа)
                        has_yabs_url = 'yabs.yandex.ru' in (ad_data['url'] or '').lower()
                        
                        # Проверяем data-cid (точный признак Яндекс.Директ)
                        has_data_cid = item.get_attribute('data-cid') is not None
                        
                        # Ищем метку "Промо" (должна быть ТОЛЬКО в первой строке и ТОЧНО!)
                        has_promo_label = False
                        text_lines = item_text.split('\n')
                        if text_lines and len(text_lines) > 0:
                            first_line = text_lines[0].strip()
                            # Проверяем ТОЧНОЕ совпадение
                            if first_line == 'Промо' or first_line == 'промо':
                                has_promo_label = True
                        
                        # Ищем метку "Реклама" (должна быть ТОЛЬКО в первой строке и ТОЧНО!)
                        has_ad_label = False
                        if text_lines and len(text_lines) > 0:
                            first_line = text_lines[0].strip()
                            # Проверяем ТОЧНОЕ совпадение
                            if first_line == 'Реклама' or first_line == 'реклама':
                                has_ad_label = True
                        
                        # Проверяем ТОЛЬКО супер-специфичные классы Директа
                        direct_specific_classes = [
                            'serp-adv__',         # Рекламный блок (с подчеркиванием - точнее)
                            'ad-snippet',         # Рекламный сниппет
                            'ya-direct'           # Явный признак Яндекс.Директ
                        ]
                        has_specific_class = any(cls in item_class.lower() for cls in direct_specific_classes)
                        
                        # МАКСИМАЛЬНО СТРОГОЕ определение: только явные признаки Директа
                        # По умолчанию считаем органикой!
                        ad_type = 'Органика'
                        
                        # Отладка: проверим условия для спорных доменов
                        debug_domains = ['magput.ru', 'alean.ru']
                        should_debug = any(domain in url_to_check for domain in debug_domains)
                        
                        if should_debug:
                            print(f"\n{'='*60}")
                            print(f"DEBUG для {url_to_check}:")
                            print(f"  has_yabs_url: {has_yabs_url}")
                            print(f"  has_data_cid: {has_data_cid}")
                            print(f"  has_specific_class: {has_specific_class}")
                            print(f"  has_promo_label: {has_promo_label}")
                            print(f"  has_ad_label: {has_ad_label}")
                            print(f"  item_class: {item_class[:200]}")
                            print(f"  first_line: '{text_lines[0] if text_lines else 'None'}'")
                            print(f"  URL: {ad_data['url']}")
                        
                        # yabs.yandex.ru - САМЫЙ НАДЕЖНЫЙ признак!
                        if has_yabs_url:
                            ad_type = 'Яндекс.Директ'
                            if should_debug: print(f"  ➜ Директ по: yabs_url")
                        # data-cid - 100% признак Директа
                        elif has_data_cid:
                            ad_type = 'Яндекс.Директ'
                            if should_debug: print(f"  ➜ Директ по: data-cid")
                        # Специфичные классы
                        elif has_specific_class:
                            ad_type = 'Яндекс.Директ'
                            if should_debug: print(f"  ➜ Директ по: specific_class ({item_class[:100]})")
                        # Метки "Промо"/"Реклама" ТОЛЬКО как первая строка и ТОЧНО
                        elif has_promo_label or has_ad_label:
                            ad_type = 'Яндекс.Директ'
                            if should_debug: print(f"  ➜ Директ по: метка в первой строке")
                        
                        if should_debug:
                            print(f"  ИТОГ: {ad_type}")
                            print(f"{'='*60}\n")
                    
                    ad_data['type'] = ad_type
                    
                    # Устанавливаем позицию в зависимости от типа
                    if ad_type == 'Яндекс.Директ':
                        direct_position += 1
                        ad_data['position'] = direct_position
                    elif ad_type == 'Органика':
                        organic_position += 1
                        ad_data['position'] = organic_position
                    elif ad_type == 'Соцсети':
                        # Соцсети не учитываем в позициях, но можем добавить отдельный счетчик
                        ad_data['position'] = 0
                    
                    # Проверяем премиум-домены
                    if ad_data['url'] or ad_data['display_url']:
                        domain_to_check = ad_data['display_url'] or ad_data['url']
                        for premium_domain in app_settings.get('premium_domains', []):
                            if premium_domain.lower() in domain_to_check.lower():
                                is_premium = True
                                ad_data['is_premium'] = True
                                ad_data['type'] = f"⭐ {ad_type} [ПРЕМИУМ]"
                                break
                    
                    if not is_premium:
                        ad_data['is_premium'] = False
                    
                    # Сохраняем только если есть заголовок
                    if ad_data['title'] and len(ad_data['title']) > 3:
                        self.results.append(ad_data)
                        ads_count += 1
                        premium_mark = " ⭐ ПРЕМИУМ" if is_premium else ""
                        logger.info(f"Объявление #{ads_count}{premium_mark} [{ad_type}]: {ad_data['title'][:50]}...")
                        self.update_status(f"Найдено объявление: {ad_data['title'][:50]}...{premium_mark}")
                
                except Exception as e:
                    logger.debug(f"Ошибка обработки элемента: {e}")
                    continue
            
            logger.info(f"Обработка завершена. Добавлено объявлений: {ads_count}")
            
            # Статистика по описаниям
            if ads_count > 0:
                ads_with_desc = sum(1 for ad in self.results[-ads_count:] if ad.get('description'))
                desc_percentage = (ads_with_desc / ads_count) * 100
                logger.info(f"📊 Описания: {ads_with_desc}/{ads_count} ({desc_percentage:.1f}%)")
                
                if desc_percentage < 70:
                    logger.warning(f"⚠ Низкий процент описаний! Рекомендуется включить режим отладки.")
            
        except Exception as e:
            logger.error(f"Ошибка в _parse_yandex_all_ads: {e}", exc_info=True)
    
    def _parse_yandex_right_ads(self, query):
        """Эта функция больше не используется - объявления парсятся через _parse_yandex_all_ads"""
        pass
    
    def _parse_yandex_top_ads(self, query):
        """Эта функция больше не используется - объявления парсятся через _parse_yandex_all_ads"""
        pass
    
    def close(self):
        """Закрытие браузера"""
        if self.driver:
            self.driver.quit()
            self.update_status("Браузер закрыт")


def parse_in_background(queries, max_pages, headless, debug_descriptions=False):
    """Функция для парсинга в фоновом потоке"""
    global parsing_status, results
    
    logger.info("="*60)
    logger.info("НАЧАЛО ФОНОВОГО ПАРСИНГА")
    logger.info(f"Запросов: {len(queries)}")
    logger.info(f"Страниц на запрос: {max_pages}")
    logger.info(f"Headless режим: {headless}")
    logger.info(f"Отладка описаний: {debug_descriptions}")
    logger.info("="*60)
    
    def status_update(message):
        parsing_status['message'] = message
    
    parsing_status['is_running'] = True
    parsing_status['progress'] = 0
    parsing_status['total_ads'] = 0
    
    parser = YandexAdParser(headless=headless, status_callback=status_update, debug_descriptions=debug_descriptions)
    
    if not parser.start_browser():
        parsing_status['is_running'] = False
        parsing_status['message'] = "Ошибка запуска браузера"
        logger.error("Не удалось запустить браузер")
        return
    
    try:
        total_queries = len(queries)
        for idx, query in enumerate(queries):
            logger.info(f"Обработка запроса {idx+1}/{total_queries}: '{query}'")
            parsing_status['current_query'] = query
            parsing_status['progress'] = int((idx / total_queries) * 100)
            
            parser.parse_yandex_ads(query, max_pages=max_pages)
            time.sleep(2)
        
        results = parser.results
        parsing_status['total_ads'] = len(results)
        parsing_status['progress'] = 100
        parsing_status['message'] = f"Парсинг завершен! Найдено {len(results)} объявлений"
        
        logger.info("="*60)
        logger.info(f"ПАРСИНГ ЗАВЕРШЕН. Всего объявлений: {len(results)}")
        logger.info("="*60)
        
        # Сохраняем результаты
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'results_{timestamp}.json'
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        logger.info(f"Результаты сохранены в {filename}")
        
    except Exception as e:
        error_msg = f"Ошибка: {str(e)}"
        parsing_status['message'] = error_msg
        logger.error(error_msg, exc_info=True)
    finally:
        parser.close()
        parsing_status['is_running'] = False
        logger.info("Фоновый парсинг завершен")


@app.route('/')
def index():
    """Главная страница"""
    return render_template('index.html')


@app.route('/start_parsing', methods=['POST'])
def start_parsing():
    """Запуск парсинга"""
    global parsing_status
    
    if parsing_status['is_running']:
        return jsonify({'success': False, 'message': 'Парсинг уже выполняется'})
    
    data = request.json
    queries = [q.strip() for q in data.get('queries', '').split('\n') if q.strip()]
    max_pages = int(data.get('max_pages', 2))
    headless = data.get('headless', True)
    debug_descriptions = data.get('debug_descriptions', False)  # Новый параметр
    
    if not queries:
        return jsonify({'success': False, 'message': 'Необходимо указать хотя бы один запрос'})
    
    # Запускаем парсинг в отдельном потоке
    thread = threading.Thread(target=parse_in_background, args=(queries, max_pages, headless, debug_descriptions))
    thread.daemon = True
    thread.start()
    
    return jsonify({'success': True, 'message': 'Парсинг запущен'})


@app.route('/status')
def get_status():
    """Получение статуса парсинга"""
    return jsonify(parsing_status)


@app.route('/results')
def get_results():
    """Получение результатов"""
    return jsonify(results)


@app.route('/download_json')
def download_json():
    """Скачивание результатов в JSON"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'yandex_ads_{timestamp}.json'
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    return send_file(filename, as_attachment=True)


@app.route('/download_csv')
def download_csv():
    """Скачивание результатов в CSV с правильной структурой"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'yandex_ads_{timestamp}.csv'
    
    if results:
        # Определяем порядок колонок для удобства
        fieldnames = [
            'position',          # Позиция
            'type',              # Тип (Директ/Органика/Соцсети)
            'query',             # Поисковый запрос
            'title',             # Заголовок
            'description',       # Описание
            'preview_url',       # Превью ссылки
            'url',               # Полный URL
            'display_url',       # Отображаемый URL
            'creative_image',    # Изображение
            'platform',          # Платформа (Яндекс)
            'is_premium',        # Премиум домен (True/False)
            'timestamp'          # Время парсинга
        ]
        
        with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            
            # Записываем заголовки на русском
            f.write('Позиция;Тип;Запрос;Заголовок;Описание;Превью;URL;Отображаемый URL;Изображение;Платформа;Премиум;Время\n')
            
            # Записываем данные
            for result in results:
                row = []
                for field in fieldnames:
                    value = result.get(field, '')
                    # Очищаем значение от точек с запятой и переносов строк
                    if value:
                        value = str(value).replace(';', ',').replace('\n', ' ').replace('\r', ' ')
                    row.append(value)
                f.write(';'.join(row) + '\n')
    
    return send_file(filename, as_attachment=True, mimetype='text/csv')


@app.route('/download_xml')
def download_xml():
    """Скачивание результатов в XML с правильной структурой"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'yandex_ads_{timestamp}.xml'
    
    # Создаем корневой элемент
    root = ET.Element('yandex_ads')
    root.set('timestamp', timestamp)
    root.set('total', str(len(results)))
    
    # Группируем по типам для удобства
    types = {}
    for ad in results:
        ad_type = ad.get('type', 'Неизвестно')
        if ad_type not in types:
            types[ad_type] = []
        types[ad_type].append(ad)
    
    # Добавляем группы по типам
    for ad_type, ads in types.items():
        type_group = ET.SubElement(root, 'group')
        type_group.set('type', ad_type)
        type_group.set('count', str(len(ads)))
        
        # Добавляем каждое объявление
        for ad in ads:
            ad_elem = ET.SubElement(type_group, 'ad')
            
            # Определяем порядок полей
            field_order = [
                'position', 'query', 'title', 'description', 
                'url', 'display_url', 'creative_image', 
                'platform', 'is_premium', 'timestamp'
            ]
            
            for key in field_order:
                if key in ad:
                    child = ET.SubElement(ad_elem, key)
                    value = ad[key]
                    # Очищаем значение
                    if value is not None:
                        child.text = str(value).strip()
                    else:
                        child.text = ''
    
    # Форматируем XML для читаемости
    xml_string = ET.tostring(root, encoding='utf-8')
    parsed = minidom.parseString(xml_string)
    pretty_xml = parsed.toprettyxml(indent='  ', encoding='utf-8')
    
    # Сохраняем в файл
    with open(filename, 'wb') as f:
        f.write(pretty_xml)
    
    return send_file(filename, as_attachment=True, mimetype='application/xml')


@app.route('/download_excel')
def download_excel():
    """Скачивание результатов в Excel (XLSX) с форматированием"""
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'openpyxl не установлен. Используйте CSV экспорт.'}), 400
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'yandex_ads_{timestamp}.xlsx'
    
    # Создаем книгу и лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты парсинга"
    
    # Заголовки с русскими названиями
    headers = [
        'Позиция', 'Тип', 'Запрос', 'Заголовок', 'Описание', 
        'Превью', 'URL', 'Отображаемый URL', 'Изображение', 'Платформа', 'Премиум', 'Время'
    ]
    
    # Стиль для заголовков
    header_fill = PatternFill(start_color="89B4FA", end_color="89B4FA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем заголовки
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Настройка ширины колонок
    column_widths = [10, 15, 20, 30, 50, 35, 40, 30, 40, 10, 10, 18]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Ключи полей
    field_keys = [
        'position', 'type', 'query', 'title', 'description',
        'preview_url', 'url', 'display_url', 'creative_image', 'platform', 'is_premium', 'timestamp'
    ]
    
    # Записываем данные
    for row_idx, result in enumerate(results, start=2):
        for col_idx, key in enumerate(field_keys, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = result.get(key, '')
            
            # Форматируем значение
            if key == 'is_premium':
                value = 'Да' if value else 'Нет'
            elif key == 'position' and value:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True, color="89B4FA")
            
            cell.value = str(value) if value else ''
            
            # Перенос текста для описания
            if key in ['description', 'title']:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Замораживаем первую строку
    ws.freeze_panes = "A2"
    
    # Сохраняем файл
    wb.save(filename)
    
    return send_file(filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/settings', methods=['GET'])
def get_settings():
    """Получение текущих настроек"""
    return jsonify(app_settings)


@app.route('/settings', methods=['POST'])
def update_settings():
    """Обновление настроек"""
    global app_settings
    
    try:
        data = request.json
        
        # Обновляем премиум-домены
        if 'premium_domains' in data:
            domains = data['premium_domains']
            if isinstance(domains, str):
                # Разбиваем строку по запятым или переносам строк
                domains = [d.strip() for d in domains.replace('\n', ',').split(',') if d.strip()]
            app_settings['premium_domains'] = domains
        
        # Обновляем другие настройки
        if 'max_pages' in data:
            app_settings['max_pages'] = int(data['max_pages'])
        
        if 'headless' in data:
            app_settings['headless'] = bool(data['headless'])
        
        # Сохраняем в файл
        if save_settings():
            logger.info(f"Настройки обновлены: {len(app_settings['premium_domains'])} премиум-доменов")
            return jsonify({
                'success': True,
                'message': 'Настройки успешно сохранены',
                'settings': app_settings
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Ошибка сохранения настроек'
            }), 500
            
    except Exception as e:
        logger.error(f"Ошибка обновления настроек: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Ошибка: {str(e)}'
        }), 500


@app.route('/premium_domains', methods=['GET'])
def get_premium_domains():
    """Получение списка премиум-доменов"""
    return jsonify({
        'domains': app_settings.get('premium_domains', []),
        'count': len(app_settings.get('premium_domains', []))
    })


@app.route('/premium_domains', methods=['POST'])
def add_premium_domain():
    """Добавление премиум-домена"""
    try:
        data = request.json
        domain = data.get('domain', '').strip()
        
        if not domain:
            return jsonify({'success': False, 'message': 'Домен не указан'}), 400
        
        if domain not in app_settings['premium_domains']:
            app_settings['premium_domains'].append(domain)
            save_settings()
            
            return jsonify({
                'success': True,
                'message': f'Домен {domain} добавлен в премиум',
                'domains': app_settings['premium_domains']
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Домен {domain} уже в списке'
            }), 400
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/premium_domains/<domain>', methods=['DELETE'])
def remove_premium_domain(domain):
    """Удаление премиум-домена"""
    try:
        if domain in app_settings['premium_domains']:
            app_settings['premium_domains'].remove(domain)
            save_settings()
            
            return jsonify({
                'success': True,
                'message': f'Домен {domain} удален из премиум',
                'domains': app_settings['premium_domains']
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Домен {domain} не найден'
            }), 404
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':
    # Загружаем настройки при старте
    load_settings()
    logger.info("="*60)
    logger.info("ПАРСЕР ЯНДЕКС РЕКЛАМЫ v2.1")
    logger.info(f"Премиум-доменов загружено: {len(app_settings.get('premium_domains', []))}")
    logger.info("="*60)
    
    app.run(debug=True, host='127.0.0.1', port=5000)
